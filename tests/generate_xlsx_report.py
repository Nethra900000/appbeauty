import os
import datetime
import openpyxl
from openpyxl.styles import Font, PatternFill, Alignment, Border, Side
from openpyxl.utils import get_column_letter

def build_excel_report(test_results, output_path="E2E_Test_Report_AppBeauty.xlsx"):
    """
    Generates a professional .xlsx report with Executive Summary Dashboard and Detailed Test Cases.
    test_results is a list of dicts with keys:
    id, category, module, title, description, preconditions, steps, expected, actual, status, priority, time_sec
    """
    wb = openpyxl.Workbook()
    
    # ----------------------------------------------------
    # Styling Definitions
    # ----------------------------------------------------
    font_title = Font(name="Calibri", size=16, bold=True, color="1F4E78")
    font_subtitle = Font(name="Calibri", size=11, italic=True, color="595959")
    font_header = Font(name="Calibri", size=11, bold=True, color="FFFFFF")
    font_bold = Font(name="Calibri", size=11, bold=True, color="000000")
    font_regular = Font(name="Calibri", size=10, color="000000")
    
    # Category / KPI Fonts
    font_kpi_num = Font(name="Calibri", size=20, bold=True, color="1F4E78")
    font_kpi_label = Font(name="Calibri", size=9, bold=True, color="595959")
    
    font_pass = Font(name="Calibri", size=10, bold=True, color="276A3C")
    font_fail = Font(name="Calibri", size=10, bold=True, color="A61C1C")
    
    fill_header = PatternFill(start_color="1F4E78", end_color="1F4E78", fill_type="solid")
    fill_sub_header = PatternFill(start_color="D9E1F2", end_color="D9E1F2", fill_type="solid")
    fill_kpi_bg = PatternFill(start_color="F2F2F2", end_color="F2F2F2", fill_type="solid")
    
    fill_pass = PatternFill(start_color="E2EFDA", end_color="E2EFDA", fill_type="solid")
    fill_fail = PatternFill(start_color="FCE4D6", end_color="FCE4D6", fill_type="solid")
    
    fill_zebra = PatternFill(start_color="F9FAFB", end_color="F9FAFB", fill_type="solid")

    thin_side = Side(border_style="thin", color="D9D9D9")
    border_all = Border(left=thin_side, right=thin_side, top=thin_side, bottom=thin_side)
    border_thick_bottom = Border(bottom=Side(border_style="medium", color="1F4E78"))
    border_kpi = Border(left=thin_side, right=thin_side, top=thin_side, bottom=thin_side)

    # ----------------------------------------------------
    # TAB 1: EXECUTIVE SUMMARY
    # ----------------------------------------------------
    ws_summary = wb.active
    ws_summary.title = "Executive Summary"
    ws_summary.views.sheetView[0].showGridLines = True

    # Header Title Block
    ws_summary.merge_cells("B2:G2")
    ws_summary["B2"] = "AI BEAUTY GENIUS (APPBEAUTY) - AUTOMATED E2E TEST REPORT"
    ws_summary["B2"].font = font_title
    
    ws_summary.merge_cells("B3:G3")
    ws_summary["B3"] = f"Generated: {datetime.datetime.now().strftime('%Y-%m-%d %H:%M:%S')} | Target: https://nethra900000.github.io/appbeauty/ | GitHub Actions CI/CD"
    ws_summary["B3"].font = font_subtitle

    # Compute Statistics
    total_tests = len(test_results)
    passed_tests = sum(1 for t in test_results if t["status"].upper() == "PASS")
    failed_tests = sum(1 for t in test_results if t["status"].upper() == "FAIL")
    pass_rate = (passed_tests / total_tests * 100) if total_tests > 0 else 0
    deploy_status = "DEPLOYABLE (PASS)" if pass_rate >= 90 else "NEEDS ATTENTION"

    # KPI Summary Cards (Row 5 - 7)
    kpis = [
        ("TOTAL TEST CASES", total_tests, "B"),
        ("PASSED TESTS", passed_tests, "C"),
        ("FAILED TESTS", failed_tests, "D"),
        ("PASS RATE", f"{pass_rate:.1f}%", "E"),
        ("DEPLOYMENT STATUS", deploy_status, "F")
    ]

    for label, val, col in kpis:
        # Label cell
        cell_lbl = ws_summary[f"{col}5"]
        cell_lbl.value = label
        cell_lbl.font = font_kpi_label
        cell_lbl.alignment = Alignment(horizontal="center", vertical="center")
        cell_lbl.fill = fill_sub_header
        cell_lbl.border = border_kpi

        # Value cell
        cell_val = ws_summary[f"{col}6"]
        cell_val.value = val
        cell_val.font = font_kpi_num if col != "F" else (font_pass if "DEPLOYABLE" in str(val) else font_fail)
        cell_val.alignment = Alignment(horizontal="center", vertical="center")
        cell_val.fill = fill_kpi_bg
        cell_val.border = border_kpi

    ws_summary.row_dimensions[5].height = 20
    ws_summary.row_dimensions[6].height = 35

    # Category Breakdown Table (Row 9)
    ws_summary["B9"] = "Category Breakdown & Quality Metrics"
    ws_summary["B9"].font = Font(name="Calibri", size=13, bold=True, color="1F4E78")

    cat_headers = ["Test Category", "Total Cases", "Passed", "Failed", "Pass Rate (%)", "Status"]
    for i, h in enumerate(cat_headers, start=2):
        col_letter = get_column_letter(i)
        cell = ws_summary[f"{col_letter}10"]
        cell.value = h
        cell.font = font_header
        cell.fill = fill_header
        cell.alignment = Alignment(horizontal="center", vertical="center")

    # Aggregate stats by category
    categories = sorted(list(set(t["category"] for t in test_results)))
    current_row = 11

    for cat in categories:
        cat_total = sum(1 for t in test_results if t["category"] == cat)
        cat_pass = sum(1 for t in test_results if t["category"] == cat and t["status"].upper() == "PASS")
        cat_fail = cat_total - cat_pass
        cat_rate = (cat_pass / cat_total * 100) if cat_total > 0 else 0
        cat_status = "HEALTHY" if cat_rate >= 90 else "WARNING"

        ws_summary[f"B{current_row}"] = cat
        ws_summary[f"C{current_row}"] = cat_total
        ws_summary[f"D{current_row}"] = cat_pass
        ws_summary[f"E{current_row}"] = cat_fail
        ws_summary[f"F{current_row}"] = f"{cat_rate:.1f}%"
        ws_summary[f"G{current_row}"] = cat_status

        for c in range(2, 8):
            cell = ws_summary[f"{get_column_letter(c)}{current_row}"]
            cell.font = font_regular
            cell.border = border_all
            if c in (3, 4, 5, 6, 7):
                cell.alignment = Alignment(horizontal="center")
            if c == 7:
                cell.font = font_pass if cat_status == "HEALTHY" else font_fail

        current_row += 1

    # Total Summary Row
    ws_summary[f"B{current_row}"] = "Total Summary"
    ws_summary[f"C{current_row}"] = total_tests
    ws_summary[f"D{current_row}"] = passed_tests
    ws_summary[f"E{current_row}"] = failed_tests
    ws_summary[f"F{current_row}"] = f"{pass_rate:.1f}%"
    ws_summary[f"G{current_row}"] = "READY" if pass_rate >= 90 else "ACTION REQ"

    for c in range(2, 8):
        cell = ws_summary[f"{get_column_letter(c)}{current_row}"]
        cell.font = font_bold
        cell.fill = fill_sub_header
        cell.border = border_all
        if c in (3, 4, 5, 6, 7):
            cell.alignment = Alignment(horizontal="center")

    # Environment Info Block
    env_row = current_row + 3
    ws_summary[f"B{env_row}"] = "Test Environment Details"
    ws_summary[f"B{env_row}"].font = Font(name="Calibri", size=12, bold=True, color="1F4E78")

    env_details = [
        ("Application Name", "AI Beauty Genius (AppBeauty)"),
        ("Repository", "Nethra900000/appbeauty"),
        ("GitHub Pages URL", "https://nethra900000.github.io/appbeauty/"),
        ("Automated Framework", "Selenium WebDriver + PyTest + HTTP Requests"),
        ("Report Engine", "OpenPyXL Test Reporter v2.0"),
        ("CI/CD Pipeline", "GitHub Actions (.github/workflows/e2e-tests.yml)"),
    ]

    for idx, (k, v) in enumerate(env_details, start=env_row+1):
        ws_summary[f"B{idx}"] = k
        ws_summary[f"C{idx}"] = v
        ws_summary[f"B{idx}"].font = font_bold
        ws_summary[f"C{idx}"].font = font_regular
        ws_summary[f"B{idx}"].border = border_all
        ws_summary[f"C{idx}"].border = border_all

    # ----------------------------------------------------
    # TAB 2: DETAILED TEST CASES
    # ----------------------------------------------------
    ws_details = wb.create_sheet(title="Detailed Test Cases")
    ws_details.views.sheetView[0].showGridLines = True

    headers = [
        "Test ID", "Category", "Module / Route", "Test Case Title", 
        "Description", "Preconditions", "Test Steps", "Expected Result", 
        "Actual Result", "Status", "Priority", "Time (s)"
    ]

    for col_num, h in enumerate(headers, 1):
        cell = ws_details.cell(row=1, column=col_num)
        cell.value = h
        cell.font = font_header
        cell.fill = fill_header
        cell.alignment = Alignment(horizontal="center", vertical="center", wrap_text=True)

    ws_details.row_dimensions[1].height = 28

    for row_idx, test in enumerate(test_results, start=2):
        status = str(test["status"]).upper()
        
        row_data = [
            test["id"],
            test["category"],
            test["module"],
            test["title"],
            test["description"],
            test["preconditions"],
            test["steps"],
            test["expected"],
            test["actual"],
            status,
            test.get("priority", "P1"),
            test.get("time_sec", 0.05)
        ]

        is_even = (row_idx % 2 == 0)

        for col_num, val in enumerate(row_data, start=1):
            cell = ws_details.cell(row=row_idx, column=col_num)
            cell.value = val
            cell.font = font_regular
            cell.border = border_all

            # Formatting specific columns
            if col_num in (1, 10, 11, 12): # Center align short fields
                cell.alignment = Alignment(horizontal="center", vertical="top")
            else:
                cell.alignment = Alignment(horizontal="left", vertical="top", wrap_text=True)

            # Apply zebra striping
            if is_even and col_num != 10:
                cell.fill = fill_zebra

            # Status cell styling
            if col_num == 10:
                if status == "PASS":
                    cell.fill = fill_pass
                    cell.font = font_pass
                else:
                    cell.fill = fill_fail
                    cell.font = font_fail

        ws_details.row_dimensions[row_idx].height = 32

    # Auto-adjust column widths
    col_widths = {
        "A": 12, "B": 24, "C": 22, "D": 32, "E": 35, 
        "F": 25, "G": 35, "H": 32, "I": 32, "J": 12, "K": 10, "L": 10
    }
    for col_letter, width in col_widths.items():
        ws_details.column_dimensions[col_letter].width = width

    # Also set Summary sheet column widths
    ws_summary.column_dimensions["A"].width = 4
    ws_summary.column_dimensions["B"].width = 30
    ws_summary.column_dimensions["C"].width = 25
    ws_summary.column_dimensions["D"].width = 16
    ws_summary.column_dimensions["E"].width = 16
    ws_summary.column_dimensions["F"].width = 22
    ws_summary.column_dimensions["G"].width = 18

    # Save Workbook
    wb.save(output_path)
    print(f"[SUCCESS] Excel Test Report generated successfully: {output_path}")
    return output_path
